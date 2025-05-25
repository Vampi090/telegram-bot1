import sqlite3
import csv
import os
from datetime import datetime
import matplotlib.pyplot as plt
from models.transaction import Transaction
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATABASE_FILE = "finance_bot.db"


def get_expense_stats(user_id):
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()
        cursor.execute("""
                       SELECT category, SUM(amount)
                       FROM transactions
                       WHERE user_id = ?
                         AND type = 'витрата'
                       GROUP BY category
                       ORDER BY SUM(amount) DESC
                       """, (user_id,))

        stats_data = cursor.fetchall()
        return stats_data
    except Exception as e:
        print(f"Error getting expense stats: {e}")
        return []
    finally:
        conn.close()


def generate_expense_chart(user_id):
    try:
        data = get_expense_stats(user_id)

        if not data:
            return None

        categories, amounts = zip(*data) if len(data) > 1 else ([data[0][0]], [data[0][1]])

        amounts = [abs(amount) for amount in amounts]

        plt.figure(figsize=(10, 6))
        bars = plt.bar(categories, amounts, color=["#4CAF50", "#FF9800", "#2196F3", "#9C27B0", "#E91E63"])

        total = sum(amounts)
        for bar in bars:
            yval = bar.get_height()
            percent = (yval / total) * 100
            plt.text(bar.get_x() + bar.get_width() / 2, yval + 10, f"{percent:.1f}%", ha='center', va='bottom')

        plt.title("📊 Розподіл витрат за категоріями")
        plt.xlabel("Категорія")
        plt.ylabel("Сума (грн)")
        plt.xticks(rotation=45, ha='right')
        plt.grid(True, axis='y', linestyle='--', alpha=0.7)

        file_path = f"chart_{user_id}.png"
        plt.tight_layout()
        plt.savefig(file_path)
        plt.close()

        return file_path
    except Exception as e:
        print(f"Error generating expense chart: {e}")
        return None


def export_transactions_to_excel(user_id):
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()

        cursor.execute("""
                       SELECT timestamp, amount, category, type
                       FROM transactions
                       WHERE user_id = ?
                       ORDER BY timestamp DESC
                       """, (user_id,))

        transactions = cursor.fetchall()
        conn.close()

        if not transactions:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакції"

        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        expense_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        headers = ["Дата", "Сума", "Категорія", "Тип"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, transaction in enumerate(transactions, 2):
            date_str = transaction[0]
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                formatted_date = date_obj.strftime("%d.%m.%Y %H:%M")
            except:
                formatted_date = date_str

            ws.cell(row=row_num, column=1).value = formatted_date
            ws.cell(row=row_num, column=2).value = abs(float(transaction[1]))
            ws.cell(row=row_num, column=3).value = transaction[2]
            ws.cell(row=row_num, column=4).value = transaction[3]

            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if transaction[3].lower() == 'дохід':
                for col_num in range(1, 5):
                    ws.cell(row=row_num, column=col_num).fill = income_fill
            elif transaction[3].lower() == 'витрата':
                for col_num in range(1, 5):
                    ws.cell(row=row_num, column=col_num).fill = expense_fill

        for col_num in range(1, 5):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, len(transactions) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 4
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = "A2"

        file_path = f"transactions_{user_id}.xlsx"
        wb.save(file_path)

        return file_path
    except Exception as e:
        print(f"Error exporting transactions to Excel: {e}")
        return None


def export_transactions_to_csv(user_id):
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()

        cursor.execute("""
                       SELECT timestamp, amount, category, type
                       FROM transactions
                       WHERE user_id = ?
                       ORDER BY timestamp DESC
                       """, (user_id,))

        transactions = cursor.fetchall()
        conn.close()

        if not transactions:
            return None

        file_path = f"transactions_{user_id}.csv"

        with open(file_path, mode="w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(["Дата", "Сума", "Категорія", "Тип"])
            writer.writerows(transactions)

        return file_path
    except Exception as e:
        print(f"Error exporting transactions to CSV: {e}")
        return None


def get_transaction_report(user_id, days=30):
    try:
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()

        cursor.execute("""
                       SELECT SUM(amount)
                       FROM transactions
                       WHERE user_id = ?
                         AND type = 'дохід'
                         AND
                           date (timestamp) >= date ('now'
                           , ?)
                       """, (user_id, f'-{days} days'))

        total_income = cursor.fetchone()[0] or 0

        cursor.execute("""
                       SELECT SUM(amount)
                       FROM transactions
                       WHERE user_id = ?
                         AND type = 'витрата'
                         AND
                           date (timestamp) >= date ('now'
                           , ?)
                       """, (user_id, f'-{days} days'))

        total_expense = cursor.fetchone()[0] or 0

        cursor.execute("""
                       SELECT category, SUM(amount)
                       FROM transactions
                       WHERE user_id = ?
                         AND type = 'витрата'
                         AND
                           date (timestamp) >= date ('now'
                           , ?)
                       GROUP BY category
                       ORDER BY SUM (amount) DESC
                           LIMIT 3
                       """, (user_id, f'-{days} days'))

        top_expense_categories = cursor.fetchall()

        cursor.execute("""
                       SELECT COUNT(*)
                       FROM transactions
                       WHERE user_id = ?
                         AND
                           date (timestamp) >= date ('now'
                           , ?)
                       """, (user_id, f'-{days} days'))

        transaction_count = cursor.fetchone()[0] or 0

        conn.close()

        return {
            'total_income': total_income,
            'total_expense': total_expense,
            'balance': total_income - total_expense,
            'top_expense_categories': top_expense_categories,
            'transaction_count': transaction_count,
            'days': days
        }
    except Exception as e:
        print(f"Error generating transaction report: {e}")
        return {
            'total_income': 0,
            'total_expense': 0,
            'balance': 0,
            'top_expense_categories': [],
            'transaction_count': 0,
            'days': days
        }
